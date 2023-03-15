import openpyxl as op
import sys

def main(argv):
    wb1 = op.load_workbook(argv[1])
    wb2 = op.load_workbook(argv[2])

    for sheet_idx in range(len(wb1.sheetnames)):
        sheet1 = wb1[wb1.sheetnames[sheet_idx]]
        sheet2 = wb2[wb2.sheetnames[sheet_idx]]

        for i, row in enumerate(sheet1.iter_rows()):
            for j, cell in enumerate(row):
                cell2 = sheet2.cell(row=i+1, column=j+1)
                if(cell.value != cell2.value):
                    print('{} {} diff'.format(i+1, j+1))
        
        print('{} sheet complete.'.format(sheet_idx+1))

if __name__ == '__main__':
    main(sys.argv)