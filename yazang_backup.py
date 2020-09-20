import openpyxl
import math
PI = 3.141592653589793

class Plant():

    species = '종'
    height = '수고(m)'

    def getName(self):
        return self.species
    
    def getHeight(self):
        return self.height
    
class Tree(Plant):
    
    def __init__(self, species, height, circum):
        self.species = str(species)
        self.height = float(height)
        self.circum = circum
    
    def getCoverage(self):
    # 만약 흉고둘레가 둘 이상인 경우, 총합을 구해서 반환하도록 되어 있음. 평균값을 구해야 한다면 len(self.circum)으로 나눠야할 것임.
        if type(self.circum) == list:
            total_coverage = 0
            for x in self.circum:
                total_coverage += (x**2) / (4*PI)
            return total_coverage
        else:
            return (self.circum**2) / (4*PI)

class Shrub(Plant):

    def __init__(self, species, height, oval, number):
        self.species = str(species)
        self.height = float(height)
        self.oval = float(oval)
        self.number = number

    def getCoverage(self):
        return self.oval

    def getNumber(self):
        return self.number

class Quadrat():

    def __init__(self, name, location, GPS_N, GPS_E, slope, aspect, altitude, geo_feature, soil_depth, veget_cover):
        self.name = name
        self.location = location
        self.GPS_N = GPS_N
        self.GPS_E = GPS_E
        self.slope = slope
        self.aspect = aspect
        self.altitude = altitude
        self.geo_feature = geo_feature
        self.soil_depth = soil_depth
        self.veget_cover = veget_cover

    def getName(self):
        return self.name
    
    def getLocation(self):
        return self.location

    def getGPS(self):
        return self.GPS_N, self.GPS_E

    def getSlope(self):
        return self.slope
    
    def getAspect(self):
        return self.aspect
    
    def getAltitude(self):
        return self.altitude
    
    def getGeoFeature(self):
        return self.geo_feature
    
    def getSoilDepth(self):
        return self.soil_depth

def main():
    target_filename = input('The name of excel file to be analyzed(with no spaces) : ')
    target_file = [target_filename]
    #target_sheet = ['야장원본 1조_1']
    target_sheet = ['야장원본 1조_1', '야장원본 1조_2', '야장원본 1조_3', '야장원본 1조_4']
    yazang_dict = {}

    for file_name in target_file:
        print('Parsing from', file_name)
        wb = openpyxl.load_workbook(file_name)
        for sheet_name in target_sheet:
            sheet_data = wb[sheet_name]
            yazang, veget_data_tree, veget_data_shrub = yazang_parse(sheet_data)
            yazang_dict[yazang] = [veget_data_tree, veget_data_shrub]
    
    yazang_analysis(yazang_dict)

def yazang_parse(sheet_data):
    # Yazang data parsing

    name = sheet_data['C4'].value
    location = sheet_data['C5'].value
    GPS = sheet_data['C7'].value
    GPS_N, GPS_E = GPS.split(', ') # GPS 엑셀 형식 조심! (ex. N: 37°28′0.14″, E: 126°57′51.90″)
    GPS_N = GPS_N.lstrip('N: ')
    GPS_E = GPS_E.lstrip('E: ')
    slope = sheet_data['F6'].value
    aspect = sheet_data['C6'].value # 방위
    altitude = sheet_data['F5'].value
    geo_feature = sheet_data['C8'].value

    # 토양 데이터 시작점 찾기
    i = 12
    while True:
        i += 1
        if sheet_data['B' + str(i)].value == '토양 채집1':
            i += 3
            break
    
    # 토양 데이터 엑셀 형식 조심 (ex. ' 1. 24')
    soil_data = []
    for j in range(i, i+5):
        n = j - i + 1
        soil_data.append(float(sheet_data['B'+str(j)].value.strip()[3:]))
        #print(soil_data[n-1])
    
    soil_depth = 0
    for depth in soil_data:
        soil_depth += depth
    soil_depth /= 5 # 땅 깊이 형식 조심

    # 식피율 데이터는 'J'열에 들어 있어야 함
    veget_cover = [sheet_data['J5'].value, sheet_data['J6'].value, sheet_data['J7'].value, sheet_data['J8'].value]

    # Yazang instance creation
    yazang = Quadrat(name, location, GPS_N, GPS_E, slope, aspect, altitude, geo_feature, soil_depth, veget_cover)
    print('\n'+name, location, GPS_N, GPS_E, slope, aspect, altitude, geo_feature, soil_depth, veget_cover, sep='\t')

    # Tree / Subtree vegetation
    print('\n### Tree & Subtree data parsing start ###\n')
    veget_data_tree = []
    next_row = 'B'

    while sheet_data[next_row + '11'].value is None or not sheet_data[next_row + '11'].value.startswith('관목'):
        i = 13
        while sheet_data[next_row + str(i)].value is not None:
            j = str(i)
            species = sheet_data[next_row + j].value
            height = sheet_data[chr(ord(next_row)+1) + j].value
            circum = sheet_data[chr(ord(next_row)+2) + j].value
            # For multiple circums
            if type(circum) == str:
                circum = list(map(float, circum.split(', ')))

            newTree = Tree(species, height, circum)
            veget_data_tree.append(newTree)
            print('New Tree:', species, height, circum)
            i += 1
        next_row = chr(ord(next_row)+3)

    # Shrub vegetation
    if sheet_data[next_row + '11'].value.startswith('관목'):
        print('\n### Shrub data parsing start ###\n')
    veget_data_shrub = []

    while sheet_data[next_row + '12'].value is not None:
        i = 13
        while sheet_data[next_row + str(i)].value is not None:
            j = str(i)
            species = sheet_data[next_row + j].value
            height = sheet_data[chr(ord(next_row)+1) + j].value
            oval = list(map(float, sheet_data[chr(ord(next_row)+2) + j].value.split(', ')))
            oval = oval[0] * oval[1] * PI / 4
            number = sheet_data[chr(ord(next_row)+3) + j].value
            # For the case if '개체수' cell is blank
            if number is None:
                number = 1
            newShrub = Shrub(species, height, oval, number)
            veget_data_shrub.append(newShrub)
            print('New Shrub:', species, height, format(oval, '.2f'), number)
            i += 1
        next_row = chr(ord(next_row)+4)

    print()
    return yazang, veget_data_tree, veget_data_shrub

def yazang_analysis(yazang_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    start_column = 5

    #i = 0
    for yazang_instance in yazang_dict.keys():
        #i += 1
        #calc_file = open('calc_' + str(i) + '.txt', 'w')

        tree_dict = appearance_analysis(yazang_dict[yazang_instance][0])
        shrub_dict = appearance_analysis(yazang_dict[yazang_instance][1])
        tree_density_dict, shrub_density_dict = density_analysis(tree_dict, shrub_dict)
        tree_coverage_dict, shrub_coverage_dict = coverage_analysis(tree_dict, shrub_dict)

        tree_relative_density_dict = make_relatively(tree_density_dict.copy())
        tree_relative_coverage_dict = make_relatively(tree_coverage_dict.copy())
        shrub_relative_density_dict = make_relatively(shrub_density_dict.copy())
        shrub_relative_coverage_dict = make_relatively(shrub_coverage_dict.copy())

        tree_importance_dict = make_importance(tree_relative_density_dict, tree_relative_coverage_dict)
        shrub_importance_dict = make_importance(shrub_relative_density_dict, shrub_relative_coverage_dict)

        next_column = yazang_output(wb, ws, start_column, yazang_instance, tree_density_dict, tree_coverage_dict, tree_relative_density_dict, tree_relative_coverage_dict, tree_importance_dict, 
        shrub_density_dict, shrub_coverage_dict, shrub_relative_density_dict, shrub_relative_coverage_dict, shrub_importance_dict)
        start_column = next_column + 1

    name = list(yazang_dict.keys())[0].getLocation()
    wb.save(name + '_testoutput.xlsx')

def appearance_analysis(species_data):
    # Species appearance analysis
 
    species_appear_set = set()
    for species_instance in species_data:
        species_appear_set.add(species_instance.getName())

    species_dict = {}
    for species_name in species_appear_set:
        for species_instance in species_data:
            if species_name == species_instance.getName():
                if species_name not in species_dict:
                    species_dict[species_name] = [species_instance]
                else:
                    species_dict[species_name].append(species_instance)

    return species_dict

def density_analysis(tree_dict, shrub_dict):
    tree_density_dict = {}
    shrub_density_dict = {}

    for tree_name in tree_dict.keys():
        tree_density_dict[tree_name] = len(tree_dict[tree_name]) / 0.01

    for shrub_name in shrub_dict.keys():
        total_shrub_number = 0
        for shrub_instance in shrub_dict[shrub_name]:
            total_shrub_number += shrub_instance.getNumber()
        shrub_density_dict[shrub_name] = total_shrub_number / 0.01

    return tree_density_dict, shrub_density_dict

def coverage_analysis(tree_dict, shrub_dict):
    tree_coverage_dict = {}
    shrub_coverage_dict = {}

    for tree_name in tree_dict.keys():
        total_tree_coverage = 0
        for tree_instance in tree_dict[tree_name]:
           total_tree_coverage += tree_instance.getCoverage()
        tree_coverage_dict[tree_name] = total_tree_coverage
    
    for shrub_name in shrub_dict.keys():
        total_shrub_coverage = 0
        for shrub_instance in shrub_dict[shrub_name]:
            total_shrub_coverage += shrub_instance.getCoverage()
        shrub_coverage_dict[shrub_name] = total_shrub_coverage

    return tree_coverage_dict, shrub_coverage_dict

def make_relatively(value_dict):
    total_value = 0
    for value in value_dict.values():
        total_value += value

    for key in value_dict.keys():
        value_dict[key] = value_dict[key] * 100 / total_value
    
    return value_dict

def make_importance(relative_density_dict, relative_coverage_dict):
    # Importance - average? or just add?

    importance_dict = {}
    for species in relative_density_dict.keys():
        importance_dict[species] = (relative_density_dict[species] + relative_coverage_dict[species]) / 2
    
    return importance_dict

def density_to_number(tree_density_dict, shrub_density_dict):
    number_dict = {}
    for tree_name in tree_density_dict:
        number_dict[tree_name] = tree_density_dict[tree_name] / 100
    for shrub_name in shrub_density_dict:
        number_dict[shrub_name] = shrub_density_dict[shrub_name] / 100
    
    total_number = 0
    for number in number_dict.values():
        total_number += number
    
    return number_dict, total_number

def margalef(number_dict, total_number):
    return (len(number_dict.keys()) - 1) / math.log(total_number)
    
def menhinick(number_dict, total_number):
    return len(number_dict.keys()) / math.sqrt(total_number)

def simpson(number_dict, total_number):
    total_sigma = 0
    for number in number_dict.values():
        total_sigma += number*(number-1)
    
    total_sigma = total_sigma / (total_number*(total_number-1))
    return 1 - total_sigma

def shannon_wiener(number_dict, total_number):
    h = 0
    for number in number_dict.values():
        p = number / total_number
        h -= p * math.log(p)
    
    return h

def yazang_output(wb, ws, start_column, yazang_instance, tree_density_dict, tree_coverage_dict, tree_relative_density_dict, tree_relative_coverage_dict, tree_importance_dict,
shrub_density_dict, shrub_coverage_dict, shrub_relative_density_dict, shrub_relative_coverage_dict, shrub_importance_dict):
    # 1. Tree & Shrub sorting by importance
    tree_list = []
    for tree_tuple in sorted(tree_importance_dict.items(), key=lambda x: x[1], reverse=True):
        tree_list.append(tree_tuple[0])    
    shrub_list = []
    for shrub_tuple in sorted(shrub_importance_dict.items(), key=lambda x: x[1], reverse=True):
        shrub_list.append(shrub_tuple[0])

    # 2. Printing data to console
    print('\n### Tree & Subtree vegetation data ###')
    for tree_name in tree_list:
        print(tree_name, format(tree_density_dict[tree_name], '.2f'), format(tree_relative_density_dict[tree_name], '.2f'), 
        format(tree_relative_coverage_dict[tree_name], '.2f'), format(tree_importance_dict[tree_name], '.2f'), sep='\t')
    print('\n### Shrub vegetation data ###')
    for shrub_name in shrub_list:
        print(shrub_name, format(shrub_density_dict[shrub_name], '.2f'), format(shrub_relative_density_dict[shrub_name], '.2f'),
        format(shrub_coverage_dict[shrub_name], '.2f'), format(shrub_relative_coverage_dict[shrub_name], '.2f'), 
        format(shrub_importance_dict[shrub_name], '.2f'), sep='\t')
    print()

    number_dict, total_number = density_to_number(tree_density_dict, shrub_density_dict)

    print('Margalef', format(margalef(number_dict, total_number), '.3f'), 'Menhinick', format(menhinick(number_dict, total_number), '.3f'), 
    'Simpson', format(simpson(number_dict, total_number), '.3f'), 'Shannon-Wiener', format(shannon_wiener(number_dict, total_number), '.3f'), '\n', sep='\t')

    # Output to excel start
    # 1. Yazang data output
    ws['A' + str(start_column)] = yazang_instance.getLocation() + '_' + yazang_instance.getName()
    ws['B' + str(start_column)] = yazang_instance.getLocation()
    ws['C' + str(start_column)] = tree_list[0]
    ws['D' + str(start_column)] = shrub_list[0]
    GPS_N, GPS_E = yazang_instance.getGPS()
    ws['E' + str(start_column)] = GPS_N
    ws['F' + str(start_column)] = GPS_E
    ws['G' + str(start_column)] = yazang_instance.getSlope()
    ws['H' + str(start_column)] = yazang_instance.getAspect()
    ws['I' + str(start_column)] = yazang_instance.getAltitude()
    ws['J' + str(start_column)] = yazang_instance.getGeoFeature()
    ws['K' + str(start_column)] = yazang_instance.getSoilDepth()

    # 2. Tree / Subtree data output
    for i in range(len(tree_list)):
        tree_name = tree_list[i]
        target_column = str(i + start_column)
        ws['L' + target_column] = tree_name
        ws['M' + target_column] = tree_density_dict[tree_name]
        ws['N' + target_column] = tree_relative_density_dict[tree_name]
        ws['O' + target_column] = tree_relative_coverage_dict[tree_name]
        ws['P' + target_column] = tree_importance_dict[tree_name]
    tree_column = int(target_column)

    # 3. Shrub data output
    for i in range(len(shrub_list)):
        shrub_name = shrub_list[i]
        target_column = str(i + start_column)
        ws['Q' + target_column] = shrub_name
        ws['R' + target_column] = shrub_density_dict[shrub_name]
        ws['S' + target_column] = shrub_relative_density_dict[shrub_name]
        ws['T' + target_column] = shrub_coverage_dict[shrub_name]
        ws['U' + target_column] = shrub_relative_coverage_dict[shrub_name]
        ws['V' + target_column] = shrub_importance_dict[shrub_name]
    shrub_column = int(target_column)

    # 4. Index data output
    ws['W' + str(start_column)] = margalef(number_dict, total_number)
    ws['X' + str(start_column)] = menhinick(number_dict, total_number)
    ws['Y' + str(start_column)] = simpson(number_dict, total_number)
    ws['Z' + str(start_column)] = shannon_wiener(number_dict, total_number)

    return max(tree_column, shrub_column)

main()
